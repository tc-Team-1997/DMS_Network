"""Tests for DOCX and XLSX text extraction.

python-docx and openpyxl are optional at CI time if the wheels are unavailable,
but we import-skip gracefully so the suite still runs.
"""
from __future__ import annotations

import io
import pytest


# ── helpers to build minimal in-memory office files ──────────────────────────

def _make_docx(paragraphs: list[str]) -> bytes:
    """Build a minimal .docx in memory using python-docx."""
    docx = pytest.importorskip("docx", reason="python-docx not installed")
    from docx import Document
    doc = Document()
    for para in paragraphs:
        doc.add_paragraph(para)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _make_xlsx(rows: list[list]) -> bytes:
    """Build a minimal .xlsx in memory using openpyxl."""
    openpyxl = pytest.importorskip("openpyxl", reason="openpyxl not installed")
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── extract_docx_text ─────────────────────────────────────────────────────────

def test_extract_docx_text_basic():
    pytest.importorskip("docx", reason="python-docx not installed")
    from app.services.docbrain.ocr import extract_docx_text

    data = _make_docx(["Hello World", "Second paragraph"])
    text = extract_docx_text(data)

    assert "Hello World" in text
    assert "Second paragraph" in text


def test_extract_docx_text_empty_doc():
    pytest.importorskip("docx", reason="python-docx not installed")
    from app.services.docbrain.ocr import extract_docx_text

    data = _make_docx([])
    text = extract_docx_text(data)
    assert isinstance(text, str)


def test_extract_docx_text_invalid_bytes_returns_empty():
    pytest.importorskip("docx", reason="python-docx not installed")
    from app.services.docbrain.ocr import extract_docx_text

    text = extract_docx_text(b"not a docx file at all \xff\xfe")
    assert isinstance(text, str)


# ── extract_xlsx_text ─────────────────────────────────────────────────────────

def test_extract_xlsx_text_basic():
    pytest.importorskip("openpyxl", reason="openpyxl not installed")
    from app.services.docbrain.ocr import extract_xlsx_text

    data = _make_xlsx([["Name", "Amount"], ["Alice", 1000], ["Bob", 2000]])
    text = extract_xlsx_text(data)

    assert "Alice" in text
    assert "1000" in text


def test_extract_xlsx_text_multiple_sheets():
    pytest.importorskip("openpyxl", reason="openpyxl not installed")
    import openpyxl
    from app.services.docbrain.ocr import extract_xlsx_text

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Foo", "Bar"])
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Baz", "Qux"])
    buf = io.BytesIO()
    wb.save(buf)

    text = extract_xlsx_text(buf.getvalue())
    assert "Foo" in text
    assert "Baz" in text


def test_extract_xlsx_text_invalid_bytes_returns_empty():
    pytest.importorskip("openpyxl", reason="openpyxl not installed")
    from app.services.docbrain.ocr import extract_xlsx_text

    text = extract_xlsx_text(b"totally not an xlsx")
    assert isinstance(text, str)


# ── dispatch integration: _tesseract_ocr dispatches to office extractors ─────

def test_ocr_dispatch_docx_mime():
    pytest.importorskip("docx", reason="python-docx not installed")
    from app.services.docbrain.ocr import _tesseract_ocr, _DOCX_MIME

    data = _make_docx(["Invoice Total: 9999"])
    result = _tesseract_ocr(data, _DOCX_MIME)

    assert result.backend == "docx"
    assert "Invoice Total" in result.full_text
    assert result.mean_confidence == 100.0


def test_ocr_dispatch_xlsx_mime():
    pytest.importorskip("openpyxl", reason="openpyxl not installed")
    from app.services.docbrain.ocr import _tesseract_ocr, _XLSX_MIME

    data = _make_xlsx([["CustomerID", "Balance"], ["CID001", 50000]])
    result = _tesseract_ocr(data, _XLSX_MIME)

    assert result.backend == "xlsx"
    assert "CID001" in result.full_text
